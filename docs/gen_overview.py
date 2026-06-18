#!/usr/bin/env python3
# gen_overview.py — generates docs/overview.html (source for docs/overview.png)
# Copyright 2026 Jim Zucker
# SPDX-License-Identifier: Apache-2.0
#
# Eight illustrative example contracts modeled on the Zucker Annuity Tracker,
# every one normalized to a $100,000 principal ($ columns in $000s). Covers all
# distinct use cases in the real data: 0% floor, negative Hard buffer, negative
# Soft barrier; capped + uncapped; participation <100/100/>100%; Annual / Monthly
# / 4Y / 5Y / 6Y resets; SPX / NDX / RUT / worst-of; Non-Qual / IRA / ROTH; plus
# a monthly-coupon income note.
#
# Downside mechanic (see README):
#   floor == 0        -> Floor at 0%: no period loss (gains capped/participated up)
#   floor < 0, floor  -> Floor (max-loss): lose only down to |floor|%
#   floor < 0, Hard   -> Hard-buffer: absorbs first |floor|%, lose 1:1 beyond
#   floor < 0, Soft   -> Soft-buffer (barrier): protected unless breached, then full loss
#   credited gain = uncapped ? part*idx : min(cap, part*idx)
#
# Regenerate:
#   python3 docs/gen_overview.py
#   "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome" \
#     --headless --disable-gpu --force-device-scale-factor=2 --hide-scrollbars \
#     --window-size=2420,470 --screenshot=docs/overview.png docs/overview.html

import datetime, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

TODAY = datetime.date(2026, 6, 14)
PRICES = {"SPX": 7400.0, "NDX": 29600.0, "RUT": 2950.0}
UNCAPPED_SENTINEL = 9.99  # v1.0 schema: numeric cell value meaning "uncapped"

# Issuer canonicalization (must match canonicalIssuer() in models.dart).
ISSUER_CANON = {
    "aspida": "ASPIDA", "athene": "ATHENE", "aig": "AIG", "axa": "AXA",
    "symetra": "SYMETRA", "citi": "CITI", "hsbc": "HSBC", "bnp": "BNP",
    "brighthouse": "BRIGHTHOUSE",
    "natbank": "NATBANK", "nationalbankofcanada": "NATBANK",
    "natbankofcanada": "NATBANK", "nbc": "NATBANK",
}

def canonical_issuer(raw):
    key = "".join(ch for ch in raw.lower() if ch.isalnum())
    return ISSUER_CANON.get(key, raw.upper())

# Reset freq v1.0 vocab: {Inception, Annual, Monthly}. Multi-year point-to-point
# resets collapse to Inception (the strike is set at inception and held to
# maturity).
def canonical_freq(raw):
    r = raw.strip().lower()
    if r == "monthly": return "Monthly"
    if r == "annual": return "Annual"
    if r == "inception": return "Inception"
    # Legacy "4-Year"/"5-Year"/"6-Year"/"N-Year" -> Inception
    return "Inception"

def d(y, m, day): return datetime.date(y, m, day)
def mdy(dt): return dt.strftime("%d-%b-%y")
def days(dt): return (dt - TODAY).days

# Each row models a real position, normalized to $100k principal.
# idx = illustrative index move for the shown period; cap=None means uncapped.
ROWS = [
    dict(pos="Aspida 12.25%-14Nov28", issuer="Aspida", index="^GSPC",
         cap=0.1225, part=1.00, floor=0.00, soft=False, idx=0.18,
         open=d(2023,11,17), last=d(2025,11,13), mat=d(2028,11,14), nxt=d(2026,11,13),
         freq="Annual", acct="Non-Qual"),
    dict(pos="Axa 65%-18Aug27", issuer="AXA", index="^GSPC",
         cap=0.65, part=1.00, floor=-0.15, soft=False, idx=-0.22,
         open=d(2021,8,18), last=d(2021,8,18), mat=d(2027,8,18), nxt=d(2027,8,18),
         freq="6-Year", acct="Non-Qual"),
    dict(pos="Citi 15%-4Feb30 IRA", issuer="Citi", index="^GSPC",
         cap=None, part=1.02, floor=-0.15, soft=False, idx=0.30,
         open=d(2025,12,31), last=d(2025,12,31), mat=d(2030,2,4), nxt=d(2030,2,4),
         freq="5-Year", acct="IRA"),
    dict(pos="HSBC 92.25%-20Oct30", issuer="HSBC", index="^NDX",
         cap=None, part=0.9225, floor=-0.15, soft=False, idx=0.40,
         open=d(2025,10,8), last=d(2025,10,3), mat=d(2030,10,20), nxt=d(2030,10,20),
         freq="5-Year", acct="IRA"),
    dict(pos="BNP 30%-6Jan31", issuer="BNP", index="^GSPC",
         cap=None, part=1.05, floor=-0.30, soft=True, idx=-0.35,
         open=d(2025,12,31), last=d(2025,12,31), mat=d(2031,1,6), nxt=d(2031,1,6),
         freq="5-Year", acct="ROTH"),
    dict(pos="NatBank 13.25%-16Apr29", issuer="Nat. Bank of Canada", index="SPX/NDX/RUT",
         cap=0.1325, part=1.00, floor=-0.30, soft=True, idx=0.0847,
         open=d(2026,4,16), last=d(2026,5,16), mat=d(2029,4,16), nxt=d(2026,6,16),
         freq="Monthly", acct="Non-Qual",
         note=True, realized=1.10, proj=0.0112, strike=6583.0,
         ndx_strike=27290.0, rut_strike=2719.0),  # income note: monthly coupon
    dict(pos="Axa 100%-20May32", issuer="AXA", index="^NDX",
         cap=1.00, part=1.00, floor=-0.20, soft=False, idx=-0.15,
         open=d(2026,5,20), last=d(2026,5,20), mat=d(2032,5,20), nxt=d(2032,5,20),
         freq="6-Year", acct="IRA"),
    dict(pos="Citi 15%-1Dec29 ROTH", issuer="Citi", index="^GSPC",
         cap=None, part=1.00, floor=-0.15, soft=False, idx=0.12,
         open=d(2025,11,28), last=d(2025,11,28), mat=d(2029,12,1), nxt=d(2029,12,1),
         freq="4-Year", acct="ROTH"),
    dict(pos="Marex 10%-12Jun31", issuer="Marex", index="^RUT",
         cap=0.20, part=1.00, floor=-0.10, floortype="floor", soft=False, idx=-0.18,
         open=d(2026,6,12), last=d(2026,6,12), mat=d(2031,6,12), nxt=d(2031,6,12),
         freq="5-Year", acct="Non-Qual"),  # max-loss Floor: −18% move clamps to −10%
]

def credited(idx, cap, part, floor, soft, floortype=None):
    if idx >= 0:
        up = part * idx
        return up if cap is None else min(cap, up)
    if floor == 0:                 # true 0% floor
        return 0.0
    if floortype == "floor":       # max-loss floor: lose only down to the floor
        return max(idx, floor)
    if soft:                       # barrier: protected unless breached
        return 0.0 if idx >= floor else idx
    return min(0.0, idx - floor)   # hard buffer: absorb first |floor|


# Protection class/label for a row: Protected (floor 0), Floor (max-loss),
# Soft (barrier), or Hard (buffer). Single source for the HTML pill + xlsx cell.
def prot_of(r):
    if r["floor"] == 0:
        return "abs", "Floor"      # 0% floor = no loss (blue pill)
    if r.get("floortype") == "floor":
        return "floor", "Floor"    # negative max-loss floor (red pill)
    if r["soft"]:
        return "soft", "Soft-buffer"
    return "hard", "Hard-buffer"

def pct(x, plus=True):
    s = f"{x*100:,.2f}%"
    return ("+" + s) if (plus and x > 0) else s

def money(x):
    return ("-$" if x < 0 else "$") + f"{abs(x):,.2f}"

def base_index(idx_name):
    if idx_name.startswith("worst") or "/" in idx_name: return "SPX"
    if idx_name == "^GSPC": return "SPX"
    if idx_name == "^NDX": return "NDX"
    if idx_name == "^RUT": return "RUT"
    return idx_name  # short name already (SPX/NDX/RUT)

# ---- compute derived values ----
for r in ROWS:
    # Issuer canonicalized to the v1.0 uppercase short form.
    r["issuer"] = canonical_issuer(r["issuer"])
    # Reset Freq canonicalized to v1.0 vocab {Inception, Annual, Monthly}.
    r["freq"] = canonical_freq(r["freq"])
    # Position is computed (matches the app): {ISSUER}-{|floor|%}-{maturity ddMMMyy}
    r["pos"] = f"{r['issuer']}-{abs(r['floor']) * 100:g}%-{r['mat']:%d%b%y}"
    if r.get("note"):
        # Income-note coupon = annual cap / 12 (matches the app's couponRate).
        r["proj_gain"] = (r["cap"] or 0) / 12
        r["realized_v"] = r["realized"]
    else:
        r["proj_gain"] = credited(
            r["idx"], r["cap"], r["part"], r["floor"], r["soft"], r.get("floortype"))
        r["realized_v"] = 0.0
        r["strike"] = PRICES[base_index(r["index"])] / (1 + r["idx"])
    # Matches the tracker: realized is reinvested into the base, so the payoff
    # applies to (initial + realized).
    base = 100.0 + r["realized_v"]
    r["proj_value"] = base * (1 + r["proj_gain"])
    r["proj_gain_dollars"] = base * r["proj_gain"]

tot_init = 100.0 * len(ROWS)
tot_real = sum(r["realized_v"] for r in ROWS)
tot_pv   = sum(r["proj_value"] for r in ROWS)
tot_pg   = sum(r["proj_gain_dollars"] for r in ROWS)

# ---- emit HTML ----
ACCT = {"Non-Qual": "nq", "IRA": "ira", "ROTH": "roth"}

def cell_floor(r):
    return "0.00%" if r["floor"] == 0 else pct(r["floor"], plus=False)

def cell_cap(r):
    if r.get("note"): return pct(r["cap"], plus=False) + " cpn"
    return "Uncapped" if r["cap"] is None else pct(r["cap"], plus=False)

rows_html = []
for r in ROWS:
    pc = "pos" if r["proj_gain"] > 0 else ("neg" if r["proj_gain"] < 0 else "")
    ic = "pos" if r["idx"] > 0 else ("neg" if r["idx"] < 0 else "")
    prot, prot_lbl = prot_of(r)
    rows_html.append(f"""      <tr>
        <td class="l">{r['issuer']}</td>
        <td class="c"><span class="pill {ACCT[r['acct']]}">{r['acct']}</span></td>
        <td class="c">{r['index']}</td>
        <td class="c"><span class="pill {prot}">{prot_lbl}</span></td>
        <td>$100.00</td><td>{money(r['realized_v'])}</td>
        <td>{money(r['proj_value'])}</td><td class="{pc}">{money(r['proj_gain_dollars'])}</td>
        <td class="{pc}">{pct(r['proj_gain'])}</td><td class="{ic}">{pct(r['idx'])}</td>
        <td class="c">{mdy(r['nxt'])}</td><td>{days(r['nxt']):,}</td>
        <td class="c">{mdy(r['mat'])}</td><td>{days(r['mat']):,}</td>
        <td>{cell_cap(r)}</td><td>{r['part']*100:,.2f}%</td><td>{cell_floor(r)}</td>
        <td>{r['strike']:,.2f}</td>
        <td class="c">{r['freq']}</td>
        <td class="c">{mdy(r['open'])}</td><td class="c">{mdy(r['last'])}</td>
      </tr>""")

HTML = f"""<!--
  overview.html — GENERATED by docs/gen_overview.py. Do not edit by hand.
  Copyright 2026 Jim Zucker
  SPDX-License-Identifier: Apache-2.0
-->
<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<style>
  :root {{ --line:#d7dbe0; --head:#1f3a5f; --headtx:#ffffff; --pos:#0a7d28; --neg:#b00020; --mut:#5b6470; }}
  * {{ box-sizing: border-box; }}
  body {{ margin: 0; padding: 20px; background:#ffffff;
         font-family: -apple-system, "Helvetica Neue", Arial, sans-serif; color:#16202b; }}
  .title {{ font-size: 19px; font-weight: 700; margin: 0 0 2px; }}
  .sub {{ font-size: 12px; color: var(--mut); margin: 0 0 12px; }}
  table {{ border-collapse: collapse; width: 2320px; font-size: 11px; }}
  thead th {{ background: var(--head); color: var(--headtx); font-weight: 600;
             padding: 6px 7px; text-align: right; vertical-align: bottom; line-height: 1.15; }}
  thead th.l {{ text-align: left; }}
  thead th.c {{ text-align: center; }}
  tbody td {{ padding: 6px 7px; text-align: right; border-bottom: 1px solid var(--line); white-space: nowrap; }}
  tbody td.l {{ text-align: left; }}
  tbody td.c {{ text-align: center; color:#33404d; }}
  tbody tr:nth-child(even) {{ background: #f5f7fa; }}
  .pill {{ display:inline-block; padding:1px 7px; border-radius:9px; font-size:10.5px; font-weight:600; }}
  .hard {{ background:#eaf7ec; color:#0a7d28; }}
  .abs {{ background:#e6efff; color:#1f3a5f; }}
  .soft {{ background:#fff3e0; color:#b26a00; }}
  .floor {{ background:#fdeaea; color:#b00020; }}
  .nq  {{ background:#fff8e1; color:#8a6d00; font-weight:600; }}
  .ira {{ background:#e6efff; color:#1f3a5f; font-weight:600; }}
  .roth{{ background:#eaf7ec; color:#0a7d28; font-weight:600; }}
  .pos {{ color: var(--pos); font-weight:600; }}
  .neg {{ color: var(--neg); font-weight:600; }}
  tfoot td {{ padding: 8px 7px; font-weight: 700; border-top: 2px solid var(--head); background:#eef1f5; text-align:right; }}
  tfoot td.l {{ text-align:left; }}
</style>
</head>
<body>
  <div class="title">Zucker Annuity Tracker &mdash; Example Contracts</div>
  <div class="sub">Nine illustrative structured products modeled on real holdings, each at a <b>$100,000</b> principal ($ columns in $000s). Floor types: <b>Floor</b> (max loss — lose only down to the floor; 0% = no loss), <b>Hard-buffer</b> (absorbs first |floor|, lose beyond), <b>Soft-buffer</b> (barrier — full loss if breached). Updated {TODAY:%d-%b-%y} &middot; illustrative prices: SPX 7,400 &nbsp; NDX 29,600 &nbsp; RUT 2,950.</div>
  <table>
    <thead>
      <tr>
        <th class="l">Issuer</th>
        <th class="c">Type</th>
        <th class="c">Index</th>
        <th class="c">Floor<br>Type</th>
        <th>Initial<br>($000)</th>
        <th>Realized<br>($000)</th>
        <th>Proj Value<br>@ Reset ($000)</th>
        <th>Proj $ Gain<br>@ Reset ($000)</th>
        <th>Proj Gain<br>@ Reset</th>
        <th>Index<br>Gain %</th>
        <th class="c">Next<br>Reset</th>
        <th>Days to<br>Reset</th>
        <th class="c">Maturity</th>
        <th>Days to<br>Maturity</th>
        <th>CAP</th>
        <th>Part.</th>
        <th>Floor</th>
        <th>Strike</th>
        <th class="c">Reset<br>Freq</th>
        <th class="c">Open</th>
        <th class="c">Last<br>Reset</th>
      </tr>
    </thead>
    <tbody>
{chr(10).join(rows_html)}
    </tbody>
    <tfoot>
      <tr>
        <td class="l" colspan="4">Totals &mdash; {len(ROWS)} contracts</td>
        <td>{money(tot_init)}</td><td>{money(tot_real)}</td><td>{money(tot_pv)}</td><td>{money(tot_pg)}</td>
        <td colspan="13"></td>
      </tr>
    </tfoot>
  </table>
</body>
</html>
"""

out = os.path.join(os.path.dirname(__file__), "overview.html")
with open(out, "w") as f:
    f.write(HTML)


# ---- emit the canonical .xlsx (Annuity Tracker schema v1.2) ----
# 24 columns (A-X), grouped Identity -> Inputs -> Outcome -> Timing -> Terms.
# Position (A) is derived. NDX_Strike / RUT_Strike (W/X) populated only for
# worst-of notes. (v1.2 lifts Inputs next to Outcome and splits the monitored
# reset/maturity dates from the static terms; the importer maps by header name,
# so v1.0/v1.1 files still load.)
HEADERS = [
    "Position",                    # A — derived, output-only
    "Issuer",                      # B — identity
    "Type",                        # C
    "Index",                       # D
    "Floor Type",                  # E — Floor | Hard-buffer | Soft-buffer
    "Initial ($000)",              # F — inputs
    "Realized ($000)",             # G
    "Proj Value @ Reset ($000)",   # H — outcome
    "Proj $ Gain @ Reset ($000)",  # I
    "Proj Gain @ Reset",           # J
    "Index Gain %",                # K
    "Next Reset",                  # L — timing (monitor)
    "Days to Reset",               # M
    "Maturity",                    # N
    "Days to Maturity",            # O
    "CAP",                         # P — terms (static); 9.99 = uncapped sentinel
    "Part.",                       # Q
    "Floor",                       # R
    "Strike",                      # S
    "Reset Freq",                  # T — Inception | Annual | Monthly
    "Open",                        # U
    "Last Reset",                  # V
    "NDX_Strike",                  # W — worst-of only
    "RUT_Strike",                  # X — worst-of only
]
PCT = "0.00%"; MONEY = "$#,##0.00"; DATE = "mm/dd/yyyy"; NUM = "#,##0.00"
HEAD_FILL = PatternFill("solid", fgColor="1F3A5F")
HEAD_FONT = Font(color="FFFFFF", bold=True)


def _floor_type_label(r):
    return prot_of(r)[1]


def _row_values(r):
    return [
        r["pos"],                                          # A Position
        r["issuer"],                                       # B Issuer
        r["acct"],                                         # C Type
        r["index"],                                        # D Index
        _floor_type_label(r),                              # E Floor Type
        100.00,                                            # F Initial ($000)
        r["realized_v"],                                   # G Realized ($000)
        r["proj_value"],                                   # H Proj Value
        r["proj_gain_dollars"],                            # I Proj $ Gain
        r["proj_gain"],                                    # J Proj Gain @ Reset
        r["idx"],                                          # K Index Gain %
        r["nxt"],                                          # L Next Reset
        days(r["nxt"]),                                    # M Days to Reset
        r["mat"],                                          # N Maturity
        days(r["mat"]),                                    # O Days to Maturity
        UNCAPPED_SENTINEL if r["cap"] is None else r["cap"],  # P CAP
        r["part"],                                         # Q Part.
        r["floor"],                                        # R Floor
        round(r["strike"], 2),                             # S Strike
        r["freq"],                                         # T Reset Freq
        r["open"],                                         # U Open
        r["last"],                                         # V Last Reset
        r.get("ndx_strike"),                               # W NDX_Strike (worst-of only)
        r.get("rut_strike"),                               # X RUT_Strike (worst-of only)
    ]


def _style_sheet(ws, header_row):
    # 1-indexed v1.2 columns:
    #   J,K,P,Q,R = pct (Proj Gain @ Reset, Index Gain %, CAP, Part., Floor)
    #   S,W,X = num (Strike, NDX/RUT strikes)
    #   L,N,U,V = date (Next Reset, Maturity, Open, Last Reset)
    #   F,G,H,I = money (Initial, Realized, Proj Value, Proj $ Gain)
    for col in (10, 11, 16, 17, 18):
        for c in ws.iter_cols(min_col=col, max_col=col, min_row=header_row + 1):
            for cell in c: cell.number_format = PCT
    for col in (12, 14, 21, 22):
        for c in ws.iter_cols(min_col=col, max_col=col, min_row=header_row + 1):
            for cell in c: cell.number_format = DATE
    for col in (6, 7, 8, 9):
        for c in ws.iter_cols(min_col=col, max_col=col, min_row=header_row + 1):
            for cell in c: cell.number_format = MONEY
    for col in (19, 23, 24):
        for c in ws.iter_cols(min_col=col, max_col=col, min_row=header_row + 1):
            for cell in c: cell.number_format = NUM


def write_xlsx(path, rows, *, with_data, with_instructions):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annuity Tracker"
    ws.append([f"ZUCKER ANNUITY TRACKER — Updated {TODAY:%d-%b-%y} "
               f"(prices: SPX {PRICES['SPX']:,.2f}  NDX {PRICES['NDX']:,.0f}  RUT {PRICES['RUT']:,.2f})"])
    ws.append(["Floor Type: Floor (max loss — lose only down to the floor; 0% floor = no loss), "
               "Hard-buffer (absorbs first |floor|, lose beyond), "
               "Soft-buffer (barrier — full loss if breached) "
               "| CAP 9.99 = uncapped | $ columns in $000s"])
    ws.append(HEADERS)
    for cell in ws[3]:
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT; cell.alignment = Alignment(wrap_text=True)
    body = rows if with_data else []
    for r in body:
        ws.append(_row_values(r))
    if with_data:
        # v1.2 money totals are contiguous: F=Initial, G=Realized, H=Proj Value,
        # I=Proj $ Gain (1-indexed 6-9). All other columns blank.
        tot = [None] * len(HEADERS)
        tot[0] = "TOTAL"
        tot[5] = 100.0 * len(rows)                          # F Initial
        tot[6] = sum(r["realized_v"] for r in rows)         # G Realized
        tot[7] = sum(r["proj_value"] for r in rows)         # H Proj Value
        tot[8] = sum(r["proj_gain_dollars"] for r in rows)  # I Proj $ Gain
        ws.append(tot)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
    _style_sheet(ws, header_row=3)
    for i, _ in enumerate(HEADERS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 15
    if with_instructions:
        ins = wb.create_sheet("Instructions")
        guide = [
            ["iHaveAnnuities — Tracker template (schema v1.2)"],
            [""],
            ["Fill one row per contract on the 'Annuity Tracker' sheet. INPUT columns"],
            ["are read by the app; DERIVED columns are recomputed on import."],
            [""],
            ["Column", "Kind", "Notes"],
            ["Position", "derived", "Never edited; rewritten on export as {ISSUER}-{|floor|%}-{ddMMMyy}"],
            ["Index Gain %", "derived", "Recomputed: current level / Strike - 1"],
            ["Proj Gain @ Reset", "derived", "Recomputed payoff for the period"],
            ["CAP", "input", "Fraction, e.g. 0.1125 = 11.25%. 9.99 = uncapped sentinel"],
            ["Part.", "input", "Participation rate, e.g. 1.00 = 100% (or 0.9225, 1.05)"],
            ["Floor", "input", "<= 0. 0 = no loss (a Floor at 0%); negative = the protection level"],
            ["Floor Type", "input", "'Floor' (max loss), 'Hard-buffer' (absorbs first |floor|), or 'Soft-buffer' (barrier)"],
            ["Strike", "input", "Index level at open / last reset (SPX strike for worst-of)"],
            ["Open / Last Reset / Maturity", "input", "Dates (mm/dd/yyyy)"],
            ["Days to Maturity", "derived", "Recomputed from Maturity"],
            ["Reset Freq", "input", "'Inception' (point-to-point), 'Annual', or 'Monthly'"],
            ["Next Reset", "input", "Date of next reset"],
            ["Days to Reset", "derived", "Recomputed from Next Reset"],
            ["Initial ($000)", "input", "Principal in thousands, e.g. 100 = $100,000"],
            ["Realized ($000)", "input", "Cumulative coupons/income to date, in $000"],
            ["Proj Value / Proj $ Gain", "derived", "Recomputed"],
            ["Type", "input", "'Non-Qual', 'IRA', or 'ROTH'"],
            ["Issuer", "input", "Canonical short name, uppercase (e.g. AIG, ASPIDA, NATBANK)"],
            ["Index", "input", "'^GSPC', '^NDX', '^RUT', or 'SPX/NDX/RUT' for worst-of"],
            ["NDX_Strike / RUT_Strike", "input", "Worst-of notes only: NDX & RUT levels at open"],
        ]
        for row in guide:
            ins.append(row)
        ins["A1"].font = Font(bold=True, size=14)
        for cell in ins[6]:
            cell.font = Font(bold=True)
        ins.column_dimensions["A"].width = 28
        ins.column_dimensions["B"].width = 10
        ins.column_dimensions["C"].width = 60
    wb.save(path)
    return path


_root = os.path.dirname(os.path.dirname(__file__))
_written = []
for _dir in (os.path.join(_root, "data"), os.path.join(_root, "app", "assets")):
    os.makedirs(_dir, exist_ok=True)
    _written.append(write_xlsx(os.path.join(_dir, "example-portfolio.xlsx"),
                               ROWS, with_data=True, with_instructions=False))
    # Template: headers + Instructions + two sample rows to copy.
    _written.append(write_xlsx(os.path.join(_dir, "template.xlsx"),
                               ROWS[:2], with_data=True, with_instructions=True))

print("wrote", out)
for _p in _written:
    print("wrote", _p)
for r in ROWS:
    print(f"  {r['pos']:26s} idx={pct(r['idx']):>9s} -> proj={pct(r['proj_gain']):>9s}  PV={money(r['proj_value']):>9s}")
print(f"  TOTAL Init={money(tot_init)} Realized={money(tot_real)} PV={money(tot_pv)} Gain={money(tot_pg)}")
